import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import argparse
from pathlib import Path

# pip install openpyxl colorama
try:
    from colorama import init, Fore, Back, Style
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False
    print("💡 Install colorama for colored output: pip install colorama")


def colored_print(message: str, color: str = "white", emoji: str = "") -> None:
    """prints colored message with emoji"""
    if not COLORAMA_AVAILABLE:
        print(f"{emoji} {message}" if emoji else message)
        return

    colors = {
        "red": Fore.RED,
        "green": Fore.GREEN,
        "yellow": Fore.YELLOW,
        "blue": Fore.BLUE,
        "magenta": Fore.MAGENTA,
        "cyan": Fore.CYAN,
        "white": Fore.WHITE,
        "bright_green": Fore.LIGHTGREEN_EX,
        "bright_blue": Fore.LIGHTBLUE_EX,
        "bright_magenta": Fore.LIGHTMAGENTA_EX,
        "bright_cyan": Fore.LIGHTCYAN_EX,
    }

    color_code = colors.get(color, Fore.WHITE)
    full_message = f"{emoji} {message}" if emoji else message
    print(f"{color_code}{Style.BRIGHT}{full_message}{Style.RESET_ALL}")


def load_json_data(filepath: str) -> dict:
    """loads json data from the specified file."""
    colored_print(f"Loading data from: {filepath}", "cyan", "📂")

    try:
        with open(filepath, "r", encoding="utf-8") as file:
            data = json.load(file)
            colored_print("JSON data loaded successfully!", "bright_green", "✅")
            return data
    except FileNotFoundError:
        colored_print(f"File not found at {filepath}", "red", "❌")
        return None
    except json.JSONDecodeError:
        colored_print(f"Could not decode JSON from {filepath}", "red", "🚫")
        return None


def create_xls_from_json(data: dict, output_filename: str) -> None:
    """creates an xlsx report from the structured json data."""
    if not data or 'days' not in data:
        colored_print("Invalid or empty JSON data provided", "red", "💥")
        return

    colored_print("Creating Excel workbook...", "yellow", "📊")

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # create style for headers
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    # headers for the main table
    headers = ["Дата", "Проект", "Задача", "Время"]
    colored_print("Adding main table headers...", "blue", "📝")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # fill the main table
    current_row = 2
    all_records = []
    record_count = 0

    colored_print("Processing daily records...", "magenta", "⚡")

    for day_data in data['days']:
        date_str = day_data.get('date')
        if not date_str:
            continue  # skip days without a date

        for task_data in day_data.get('done', []):
            project_type = task_data.get('type', 'task')  # default to 'task'
            project = 'meet' if project_type == 'meeting' else 'dev'
            task_description = task_data.get('task', 'no description')
            duration = task_data.get('duration', 0)

            record = [date_str, project, task_description, float(duration)]
            all_records.append(record)
            record_count += 1

            for col, value in enumerate(record, start=1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

    colored_print(f"Processed {record_count} task records", "bright_cyan", "🎯")

    # headers for the summary table
    summary_headers = ["Дата", "Время"]
    colored_print("Adding summary table...", "blue", "📈")

    for i, header in enumerate(summary_headers):
        col_idx = 6 + i
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # calculate and fill the summary table
    daily_totals = {}
    total_time = 0
    for record in all_records:
        date = record[0]
        time = record[3]
        if date in daily_totals:
            daily_totals[date] += time
        else:
            daily_totals[date] = time
        total_time += time

    summary_row = 2
    # sort dates for the summary table
    sorted_dates = sorted(daily_totals.keys(), key=lambda d: datetime.strptime(d, '%Y-%m-%d'))

    for date in sorted_dates:
        time = daily_totals[date]
        ws.cell(row=summary_row, column=6, value=date)
        ws.cell(row=summary_row, column=7, value=time)
        summary_row += 1

    # add the "Итого" (total) row
    total_row_idx = summary_row
    ws.cell(row=total_row_idx, column=6, value="Итого")
    ws.cell(row=total_row_idx, column=7, value=total_time)
    ws.cell(row=total_row_idx, column=6).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=7).font = Font(bold=True)

    colored_print(f"Total time calculated: {total_time} hours", "bright_green", "⏰")

    # set column widths
    colored_print("Formatting columns...", "yellow", "🎨")
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10  # 'dev' or 'meet'
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    try:
        wb.save(output_filename)
        colored_print(f"Report saved to: {output_filename}", "bright_green", "🎉")
    except Exception as e:
        colored_print(f"Error saving Excel file: {e}", "red", "💀")


if __name__ == "__main__":
    colored_print("Excel Report Generator", "bright_magenta", "🌈")
    colored_print("=" * 50, "magenta")

    json_filepath = None

    parser = argparse.ArgumentParser(description="Generate an Excel report from JSON data.")
    parser.add_argument(
        "-f", "--file",
        type=str,
        required=False,
        help="Path to the JSON file containing the report data."
    )
    args = parser.parse_args()
    json_filepath = args.file

    if not json_filepath:
        raise ValueError("Please provide a valid JSON file path using the -f or --file argument.")
    else:
        colored_print(f"Using provided JSON file path: {json_filepath}", "cyan", "📂")

    report_data = load_json_data(json_filepath)

    if report_data:
        # month, year = datetime.now().strftime("%m-%Y").split('-')

        # get first day of the report data to use in the report filename
        first_day = report_data.get('days', [{}])[0].get('date', 'unknown')

        month_str, year = first_day.split('-')[:2]
        output_xls_filename = Path(reports_dir) / f"Anton_Karmanov_{month_str}_{year}_v{month}.xlsx"

        import os
        os.makedirs(reports_dir, exist_ok=True)

        colored_print("Starting report generation...", "bright_blue", "🚀")
        create_xls_from_json(report_data, output_xls_filename)
        colored_print(f"'{output_xls_filename}' created successfully!", "bright_green", "🎊")
        colored_print("All done! Have a fabulous day! ✨", "bright_magenta", "💖")
    else:
        colored_print("Report generation failed due to data loading issues", "red", "😢")

# requirements.txt:
# openpyxl>=3.0.9 # ensure you have a recent version
# colorama>=0.4.4 # for beautiful colored output
