#!/usr/bin/env python3
"""
Report formatter
Creates Excel reports from JSON report data - matches working example
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional

# pip install openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("❌ Install required packages: pip install openpyxl")
    exit(1)

logger = logging.getLogger(__name__)


class ReportFormatter:
    """Formats report data into Excel documents - matches working example"""
    
    def __init__(self, reports_dir: str = 'reports'):
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(exist_ok=True)
    
    def load_json_data(self, json_path: str) -> Optional[Dict]:
        """Load report data from JSON file"""
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.error(f"❌ Failed to load JSON data: {e}")
            return None
    
    def create_excel_report(self, data: Dict, output_path: Optional[str] = None) -> str:
        """Create Excel report from JSON data - matches working example logic"""
        if not data or 'days' not in data:
            raise ValueError("No valid report data provided")
            
        logger.info("📊 Creating Excel workbook...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        # create style for headers
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')

        # headers for the main table
        headers = ["Дата", "Проект", "Задача", "Время"]
        logger.info("📝 Adding main table headers...")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # fill the main table
        current_row = 2
        all_records = []
        record_count = 0

        logger.info("⚡ Processing daily records...")

        for day_data in data['days']:
            date = day_data['date']
            for task in day_data.get('done', []):
                # Map task type to project type
                project_type = "meet" if task.get('type') == 'meeting' else "dev"
                task_description = task.get('description', task.get('task', ''))
                duration = task.get('duration', 0)
                
                # Add to main table
                ws.cell(row=current_row, column=1, value=date)
                ws.cell(row=current_row, column=2, value=project_type)
                ws.cell(row=current_row, column=3, value=task_description)
                ws.cell(row=current_row, column=4, value=duration)
                
                # Store for summary calculation
                all_records.append({
                    'date': date,
                    'type': project_type,
                    'description': task_description,
                    'duration': duration
                })
                
                current_row += 1
                record_count += 1

        logger.info(f"🎯 Processed {record_count} task records")

        # headers for the summary table
        summary_headers = ["Дата", "Время"]
        logger.info("📈 Adding summary table...")

        for i, header in enumerate(summary_headers):
            col = 6 + i  # Start at column F
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # calculate and fill the summary table
        daily_totals = {}
        total_time = 0
        for record in all_records:
            date = record['date']
            duration = record['duration']
            if date not in daily_totals:
                daily_totals[date] = 0
            daily_totals[date] += duration
            total_time += duration

        summary_row = 2
        # sort dates for the summary table
        sorted_dates = sorted(daily_totals.keys(), key=lambda d: datetime.strptime(d, '%Y-%m-%d'))

        for date in sorted_dates:
            ws.cell(row=summary_row, column=6, value=date)
            ws.cell(row=summary_row, column=7, value=daily_totals[date])
            summary_row += 1

        # add the "Итого" (total) row
        total_row_idx = summary_row
        ws.cell(row=total_row_idx, column=6, value="Итого")
        ws.cell(row=total_row_idx, column=7, value=total_time)
        ws.cell(row=total_row_idx, column=6).font = Font(bold=True)
        ws.cell(row=total_row_idx, column=7).font = Font(bold=True)

        logger.info(f"⏰ Total time calculated: {total_time} hours")

        # set column widths
        logger.info("🎨 Formatting columns...")
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10  # 'dev' or 'meet'
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10

        # Generate output path if not provided
        output_file_path: Path
        if not output_path:
            # Extract month and year from first day
            first_day = data.get('days', [{}])[0].get('date', 'unknown')
            year, month_str = first_day.split('-')[:2]
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            output_filename = f"Anton_Karmanov_{year}_{month_str}_v{timestamp}.xlsx"
            output_file_path = self.reports_dir / output_filename
        else:
            output_file_path = Path(output_path)

        # Ensure directory exists
        output_file_path.parent.mkdir(exist_ok=True)
        
        try:
            wb.save(str(output_file_path))
            logger.info(f"✅ Excel report saved successfully: {output_file_path}")
            return str(output_file_path)
        except Exception as e:
            logger.error(f"❌ Failed to save Excel report: {e}")
            raise
